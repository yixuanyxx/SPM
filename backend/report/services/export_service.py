import os
import io
from typing import Dict, Any, List
from datetime import datetime
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl.styles import PatternFill

from models.report import ReportData, TeamReportData


class ExportService:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        
        # Custom styles
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.darkblue,
            alignment=1  # Center alignment
        )
        
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.darkblue,
            spaceAfter=12
        )

    def export_personal_report_pdf(self, report_data: ReportData) -> bytes:
        """Export personal report as PDF organized by projects"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        # Title
        title = f"Personal Report - {report_data.user_name} ({report_data.user_role.capitalize()})"
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 20))

        # Summary Overview
        story.append(Paragraph("Summary Overview", self.heading_style))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Projects', str(report_data.total_projects)],
            ['Total Tasks', str(report_data.total_tasks)],
            ['Completed Tasks', str(report_data.completed_tasks)],
            ['Completion Rate', f"{report_data.completion_percentage:.1f}%"],
            ['Overdue Tasks', str(getattr(report_data, 'overdue_tasks', 0))],
            ['Average Task Duration', f"{report_data.average_task_duration:.1f} days" if report_data.average_task_duration else "N/A"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Projects and Tasks Details
        projects_breakdown = getattr(report_data, 'projects_breakdown', [])
        if projects_breakdown:
            story.append(Paragraph("Projects and Tasks Details", self.heading_style))
            
            for project in projects_breakdown:
                project_name = project.get('project_name', 'Unknown Project')
                
                # Project Header
                story.append(Paragraph(f"Project: {project_name}", self.styles['Heading2']))
                
                # Project Summary
                project_summary = [
                    ['Project Metric', 'Value'],
                    ['Total Tasks', str(project.get('total_tasks', 0))],
                    ['Completed Tasks', str(project.get('completed_tasks', 0))],
                    ['Completion Rate', f"{project.get('completion_percentage', 0):.1f}%"],
                    ['Overdue Tasks', str(project.get('overdue_tasks', 0))],
                    ['Average Duration', f"{project.get('average_task_duration', 0):.1f} days" if project.get('average_task_duration') else "N/A"],
                    ['Projected Completion', project.get('projected_completion_date', 'N/A')]
                ]
                
                project_table = Table(project_summary, colWidths=[2*inch, 2.5*inch])
                project_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(project_table)
                story.append(Spacer(1, 15))
                
                # Tasks in this project
                task_details = project.get('task_details', [])
                if task_details:
                    story.append(Paragraph(f"Tasks in {project_name} ({len(task_details)} tasks) - ★ = You are owner/collaborator", self.styles['Heading3']))
                    
                    task_data = [['Task Name', 'Status', 'Owner', 'Collaborators', 'Due Date', 'Duration', 'Notes']]
                    
                    # Track which rows need highlighting
                    highlighted_rows = []
                    
                    for task_idx, task in enumerate(task_details):
                        # Check if user is owner or collaborator
                        user_id = report_data.user_id
                        owner_id = task.get('owner_id')
                        collaborator_ids = task.get('collaborator_ids', []) or []
                        
                        is_user_involved = (owner_id == user_id) or (user_id in collaborator_ids)
                        
                        # Get task information
                        task_name = task.get('task_name', 'Unknown')
                        if is_user_involved:
                            task_name = f"★ {task_name}"
                            highlighted_rows.append(task_idx + 1)  # +1 because row 0 is header
                        
                        status = task.get('status', 'Unknown')
                        owner = task.get('owner_name', 'Unknown')
                        collaborators_display = ", ".join(task.get('collaborators', [])) if task.get('collaborators') else "None"
                        due_date = task.get('due_date', '')[:10] if task.get('due_date') else 'N/A'
                        
                        # Duration and status notes
                        completion_days = task.get('completion_days')
                        duration_text = f"{completion_days}d" if completion_days else "Ongoing"
                        
                        notes = []
                        if task.get('is_overdue'):
                            notes.append(f"Overdue {task.get('days_overdue', 0)}d")
                        if task.get('was_completed_late'):
                            notes.append("Late completion")
                        if task.get('priority') == 'High':
                            notes.append("High priority")
                        
                        notes_text = "; ".join(notes) if notes else "On track"
                        
                        task_data.append([
                            task_name,
                            status,
                            owner,
                            collaborators_display,
                            due_date,
                            duration_text,
                            notes_text
                        ])
                    
                    task_table = Table(task_data, colWidths=[1.5*inch, 0.8*inch, 0.8*inch, 1.2*inch, 0.8*inch, 0.6*inch, 1.3*inch])
                    
                    # Build table style with highlighting for user's tasks
                    table_style = [
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 7),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP')
                    ]
                    
                    # Add yellow highlighting for rows where user is involved
                    for row_idx in highlighted_rows:
                        table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightyellow))
                    
                    task_table.setStyle(TableStyle(table_style))
                    story.append(task_table)
                    story.append(Spacer(1, 20))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

        # Project Focus & Critical Tasks - Comprehensive Information
        if report_data.project_stats:
            story.append(Paragraph("Project Portfolio & Task Distribution", self.heading_style))
            
            # Show all projects, not just top 3
            for project in report_data.project_stats:
                project_name = project.get('project_name', 'Unknown Project')
                project_status = "On Track" if project['completion_percentage'] >= 70 and project.get('overdue_percentage', 0) < 20 else \
                                "At Risk" if project.get('overdue_percentage', 0) < 40 else "Critical"
                
                story.append(Paragraph(f"{project_name} | Status: {project_status}", self.styles['Heading3']))
                
                # Comprehensive project information
                project_summary = [
                    ['Tasks Progress', f"{project['completed_tasks']}/{project['total_tasks']} tasks completed ({project['completion_percentage']:.0f}%)"],
                    ['Task Status Distribution', f"Ongoing: {project.get('in_progress_tasks', 0)}, Under Review: {project.get('under_review_tasks', 0)}"],
                    ['Overdue Tasks', f"{project.get('overdue_tasks', 0)} tasks" + (f" ({project.get('overdue_percentage', 0):.0f}%)" if project.get('overdue_tasks', 0) > 0 else " (None)")],
                    ['Late Completions', f"{project.get('late_completions', 0)} tasks"],
                    ['On-Time Completion Rate', f"{project.get('on_time_completion_rate', 0):.1f}%"],
                ]
                
                if project.get('average_task_duration'):
                    project_summary.append(['Average Task Duration', f"{project['average_task_duration']:.1f} days"])
                
                if project.get('projected_completion_date') and project['projected_completion_date'] != 'Completed':
                    project_summary.append(['Estimated Completion', project['projected_completion_date'][:10]])
                elif project['projected_completion_date'] == 'Completed':
                    project_summary.append(['Status', 'Project Completed'])
                
                # Show team members working on this project
                task_assignees = project.get('task_assignees', {})
                if task_assignees:
                    assignee_names = []
                    for assignee_info in task_assignees.values():
                        name = assignee_info.get('owner_name', 'Unknown')
                        task_count = len(assignee_info.get('tasks', []))
                        assignee_names.append(f"{name} ({task_count} tasks)")
                    
                    if len(assignee_names) <= 3:
                        assignees_text = ", ".join(assignee_names)
                    else:
                        assignees_text = ", ".join(assignee_names[:3]) + f" and {len(assignee_names) - 3} others"
                    
                    project_summary.append(['Team Members', assignees_text])
                
                proj_table = Table(project_summary, colWidths=[1.8*inch, 4.2*inch])
                proj_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightcyan),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]))
                story.append(proj_table)
                story.append(Spacer(1, 12))
        
        # Critical Tasks & Action Items - Complete list with full information
        task_details = getattr(report_data, 'task_details', [])
        critical_tasks = [task for task in task_details if 
                         task.get('is_overdue') or 
                         task.get('priority') == 'High' or 
                         task.get('status') in ['Under Review']]
        
        if critical_tasks:
            story.append(Paragraph("Action Required - All Critical Tasks (★ = You are owner/collaborator)", self.heading_style))
            
            action_data = [['Task Name', 'Project', 'Priority', 'Status', 'Due Date', 'Action Needed']]
            
            # Track which rows need highlighting
            highlighted_rows = []
            
            # Show ALL critical tasks, not limited to 8
            for task_idx, task in enumerate(critical_tasks):
                # Check if user is owner or collaborator
                user_id = report_data.user_id
                owner_id = task.get('owner_id')
                collaborator_ids = task.get('collaborator_ids', []) or []
                
                is_user_involved = (owner_id == user_id) or (user_id in collaborator_ids)
                
                action_needed = "OVERDUE!" if task.get('is_overdue') else \
                               "Review Required" if task.get('status') == 'Under Review' else \
                               "High Priority" if task.get('priority') == 'High' else ""
                               
                due_date = task.get('due_date', '')[:10] if task.get('due_date') else 'N/A'
                task_name = task.get('task_name', 'Unknown')
                if is_user_involved:
                    task_name = f"★ {task_name}"
                    highlighted_rows.append(task_idx + 1)  # +1 because row 0 is header
                
                project_name = task.get('project_name', 'No Project')
                
                action_data.append([
                    task_name,  # Don't truncate task names
                    project_name,
                    task.get('priority', 'Normal'),
                    task.get('status', 'Unknown'),
                    due_date,
                    action_needed
                ])
            
            # Adjust table column widths to prevent truncation
            action_table = Table(action_data, colWidths=[2.2*inch, 1.3*inch, 0.7*inch, 0.8*inch, 0.7*inch, 1.0*inch])
            
            # Build table style with highlighting for user's tasks
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP')
            ]
            
            # Add yellow highlighting for rows where user is involved
            for row_idx in highlighted_rows:
                table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightyellow))
            
            action_table.setStyle(TableStyle(table_style))
            story.append(action_table)
            story.append(Spacer(1, 20))

        # Comprehensive Task Details - Complete Information with No Truncation
        task_details = getattr(report_data, 'task_details', [])
        if task_details:
            story.append(Paragraph("Complete Task Portfolio & Performance Analysis (★ = You are owner/collaborator)", self.heading_style))
            
            # Group tasks by status for organized viewing
            task_groups = {}
            for task in task_details:
                status = task.get('status', 'Unknown')
                if status not in task_groups:
                    task_groups[status] = []
                task_groups[status].append(task)
            
            # Display each status group with complete information
            for status, tasks in task_groups.items():
                if tasks:
                    story.append(Paragraph(f"{status} Tasks ({len(tasks)})", self.styles['Heading3']))
                    
                    # Comprehensive task table with all relevant details - no truncation
                    task_detail_data = [['Task Name', 'Project', 'Owner', 'Collaborators', 'Duration', 'Due Date', 'Completed', 'Status Notes']]
                    
                    # Track which rows need highlighting
                    highlighted_rows = []
                    
                    for task_idx, task in enumerate(tasks):
                        # Check if user is owner or collaborator
                        user_id = report_data.user_id
                        owner_id = task.get('owner_id')
                        collaborator_ids = task.get('collaborator_ids', []) or []
                        
                        is_user_involved = (owner_id == user_id) or (user_id in collaborator_ids)
                        
                        # Get full task name - no truncation
                        task_name = task.get('task_name', 'Unknown')
                        if is_user_involved:
                            task_name = f"★ {task_name}"
                            highlighted_rows.append(task_idx + 1)  # +1 because row 0 is header
                        
                        # Get project name
                        project_name = task.get('project_name', 'No Project')
                        
                        # Get owner name
                        owner_name = task.get('owner_name', 'Unknown')
                        
                        # Get all collaborators (full list as requested)
                        collaborators_list = task.get('collaborators', [])
                        collaborators_text = "None"
                        if collaborators_list:
                            collaborators_text = ", ".join(collaborators_list)
                        
                        # Get completion duration
                        completion_days = task.get('completion_days')
                        duration_text = f"{completion_days} days" if completion_days else "Ongoing"
                        
                        # Get due date
                        due_date = task.get('due_date', '')[:10] if task.get('due_date') else 'N/A'
                        
                        # Get completion date
                        completed_date = task.get('completed_at', '')[:10] if task.get('completed_at') else 'N/A'
                        
                        # Status notes
                        status_notes = []
                        if task.get('is_overdue'):
                            days_overdue = task.get('days_overdue', 0)
                            status_notes.append(f"Overdue by {days_overdue} days")
                        if task.get('was_completed_late'):
                            status_notes.append("Completed late")
                        if task.get('priority') == 'High':
                            status_notes.append("High priority")
                        
                        status_notes_text = "; ".join(status_notes) if status_notes else "On track"
                        
                        task_detail_data.append([
                            task_name,
                            project_name,
                            owner_name,
                            collaborators_text,
                            duration_text,
                            due_date,
                            completed_date,
                            status_notes_text
                        ])
                    
                    # Use wider table with proper column widths to prevent truncation
                    task_table = Table(task_detail_data, colWidths=[1.5*inch, 1.0*inch, 0.8*inch, 1.2*inch, 0.7*inch, 0.8*inch, 0.8*inch, 1.2*inch])
                    
                    # Build table style with highlighting for user's tasks
                    table_style = [
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 7),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP')
                    ]
                    
                    # Add yellow highlighting for rows where user is involved
                    for row_idx in highlighted_rows:
                        table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightyellow))
                    
                    task_table.setStyle(TableStyle(table_style))
                    story.append(task_table)
                    story.append(Spacer(1, 15))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()


    def export_team_report_pdf(self, manager_report: ReportData = None, team_report: TeamReportData = None, detailed_workload: Dict[str, Any] = None) -> bytes:
        """Export team report with overview section and detailed member sections showing all project tasks with highlighting"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        # ========== SECTION 1: OVERVIEW ==========
        # Title - differentiate between team and department reports
        if team_report.dept_name and not team_report.team_name:
            title = f"Department Report: {team_report.dept_name}"
            overview_title = "Department Overview"
        else:
            title = f"Team Report: {team_report.team_name or team_report.dept_name}"
            overview_title = "Team Overview"
        
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 20))

        # Team/Department Overview
        story.append(Paragraph(overview_title, self.heading_style))
        
        # Determine the label for team/department row
        entity_label = 'Department' if (team_report.dept_name and not team_report.team_name) else 'Team'
        entity_value = team_report.dept_name if (team_report.dept_name and not team_report.team_name) else (team_report.team_name or 'Unknown')
        size_label = 'Department Size' if (team_report.dept_name and not team_report.team_name) else 'Team Size'
        
        team_overview = [
            ['Metric', 'Value'],
            [entity_label, entity_value],
            [size_label, str(len(team_report.member_reports))],
            ['Total Projects', str(team_report.total_team_projects)],
            ['Total Tasks', str(team_report.total_team_tasks)],
            ['Team Completion Rate', f"{team_report.team_completion_percentage:.1f}%"],
            ['Team Overdue Rate', f"{getattr(team_report, 'team_overdue_percentage', 0):.1f}%"],
            ['Avg Task Duration', f"{team_report.team_average_task_duration:.1f} days" if team_report.team_average_task_duration else "N/A"]
        ]
        
        overview_table = Table(team_overview, colWidths=[2.5*inch, 2.5*inch])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(overview_table)
        story.append(Spacer(1, 30))
        
        # ========== MEMBER SECTIONS (One per member, grouped by team for departments) ==========
        # Get enriched project stats from team report
        team_project_stats = getattr(team_report, 'team_project_stats', [])
        
        # For department reports, group members by team
        is_department_report = team_report.dept_name and not team_report.team_name
        
        if is_department_report:
            # Group members by team
            teams_dict = {}
            for member in team_report.member_reports:
                team_id = getattr(member, 'team_id', None)
                team_name = getattr(member, 'team_name', None)
                
                # Handle missing team info
                if not team_id:
                    team_id = 'no_team'
                    team_name = 'No Team Assigned'
                elif not team_name:
                    # If team_id exists but team_name is None, use fallback
                    team_name = f'Team {team_id}'
                
                if team_id not in teams_dict:
                    teams_dict[team_id] = {
                        'team_name': team_name,
                        'members': []
                    }
                teams_dict[team_id]['members'].append(member)
            
            # Process each team with its members
            for team_idx, (team_id, team_data) in enumerate(teams_dict.items()):
                if team_idx > 0:
                    story.append(PageBreak())
                
                # Team header for department reports
                team_header = Paragraph(f"TEAM: {team_data['team_name']}", 
                                       ParagraphStyle('TeamHeader',
                                                    parent=getSampleStyleSheet()['Heading1'],
                                                    fontSize=18,
                                                    textColor=colors.darkred,
                                                    spaceAfter=20))
                story.append(team_header)
                story.append(Spacer(1, 10))
                
                # Process members in this team
                for local_member_idx, member in enumerate(team_data['members'], 1):
                    if local_member_idx > 1:
                        story.append(Spacer(1, 30))  # Space between members in same team
                    
                    # Process this member (inline member processing)
                    self._add_member_section_to_pdf(story, member, local_member_idx, team_project_stats, is_department_report)
        else:
            # Regular team report - process members directly
            if not team_report.member_reports:
                # No members in team
                story.append(Paragraph("No members in this team", 
                                     ParagraphStyle('NoMembers',
                                                  parent=getSampleStyleSheet()['Normal'],
                                                  fontSize=11,
                                                  textColor=colors.grey,
                                                  spaceAfter=15)))
            
            for member_idx, member in enumerate(team_report.member_reports, 1):
                if member_idx > 1:
                    story.append(PageBreak())  # New page for each member in team reports
                
                # Process this member
                self._add_member_section_to_pdf(story, member, member_idx, team_project_stats, is_department_report)
        
        # Build and return PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _add_member_section_to_pdf(self, story, member, member_idx, team_project_stats, is_department_report):
        """Add a member section to the PDF story"""
        # Member Header
        member_title = Paragraph(f"Member {member_idx}: {member.user_name}", 
                                ParagraphStyle('MemberTitle',
                                             parent=getSampleStyleSheet()['Heading1'],
                                             fontSize=16,
                                             textColor=colors.darkblue,
                                             spaceAfter=15))
        story.append(member_title)
        
        # Member Performance Summary
        member_info = [
            ['Metric', 'Value'],
            ['Role', member.user_role.capitalize() if member.user_role else 'Unknown'],
        ]
        
        # Add team info for department reports
        if is_department_report:
            team_name = getattr(member, 'team_name', None)
            if not team_name:
                team_id = getattr(member, 'team_id', None)
                team_name = f'Team {team_id}' if team_id else 'No Team Assigned'
            member_info.append(['Team', team_name])
        
        member_info.extend([
            ['Total Projects', member.total_projects],
            ['Total Tasks', member.total_tasks],
            ['Completed Tasks', member.completed_tasks],
            ['Completion Rate', f"{member.completion_percentage:.1f}%"],
            ['Overdue Tasks', getattr(member, 'overdue_tasks', 0)],
            ['Average Task Duration', f"{member.average_task_duration:.1f} days" if member.average_task_duration else "N/A"]
        ])
        
        member_info_table = Table(member_info, colWidths=[2.5*inch, 2.5*inch])
        member_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(member_info_table)
        story.append(Spacer(1, 20))
        
        # Projects this member is involved in
        member_user_id = member.user_id
        
        # Ensure member_user_id is an integer for comparison
        try:
            member_user_id = int(member_user_id)
        except (ValueError, TypeError):
            pass  # Keep original type if conversion fails
        
        projects_found = 0
        for project in team_project_stats:
            project_id = project.get('project_id')
            member_involvement = project.get('member_involvement', {})
            
            # Check if this member is involved in the project
            if member_user_id in member_involvement:
                projects_found += 1
                involved_task_ids = member_involvement[member_user_id]['involved_tasks']
                
                # Project Header
                project_title = Paragraph(f"PROJECT: {project.get('project_name', 'Unknown Project')}", 
                                         ParagraphStyle('ProjectTitle',
                                                      parent=getSampleStyleSheet()['Heading2'],
                                                      fontSize=12,
                                                      textColor=colors.darkgreen,
                                                      spaceAfter=10))
                story.append(project_title)
                
                # Project-level metrics
                project_metrics = [
                    ['Metric', 'Value'],
                    ['Total Tasks', project.get('total_tasks', 0)],
                    ['Completed Tasks', project.get('completed_tasks', 0)],
                    ['Ongoing Tasks', project.get('in_progress_tasks', 0)],
                    ['Under Review Tasks', project.get('under_review_tasks', 0)],
                    ['Overdue Tasks', project.get('overdue_tasks', 0)],
                    ['Completion Rate', f"{project.get('completion_percentage', 0):.1f}%"],
                    ['Average Task Duration', f"{project.get('average_task_duration'):.1f} days" if project.get('average_task_duration') else "N/A"],
                    ['Projected Completion', project.get('projected_completion_date') if project.get('projected_completion_date') else "Completion date cannot be projected"]
                ]
                
                project_metrics_table = Table(project_metrics, colWidths=[2*inch, 2*inch])
                project_metrics_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(project_metrics_table)
                story.append(Spacer(1, 12))
                
                # ALL Tasks in this project (with highlighting)
                all_project_tasks = project.get('all_tasks', [])
                if all_project_tasks:
                    tasks_header = Paragraph("All Tasks in Project (★ = Member is owner/collaborator)", 
                                            ParagraphStyle('TasksHeader',
                                                         parent=getSampleStyleSheet()['Heading3'],
                                                         fontSize=10,
                                                         spaceAfter=8))
                    story.append(tasks_header)
                    
                    # Task table
                    task_data = [['Task Name', 'Status', 'Priority', 'Owner', 'Collaborators', 'Due Date', 'Duration']]
                    
                    # Track which rows need highlighting
                    highlighted_rows = []
                    
                    for task_idx, task in enumerate(all_project_tasks):
                        task_id = task.get('id')
                        is_member_involved = task_id in involved_task_ids
                        
                        # Get task details - NO TRUNCATION
                        task_name = task.get('task_name', 'Unknown Task')
                        if is_member_involved:
                            task_name = f"★ {task_name}"
                            highlighted_rows.append(task_idx + 1)  # +1 because row 0 is header
                        
                        # Get owner name - NO TRUNCATION
                        owner_name = task.get('owner_name', 'Unknown')
                        
                        # Get collaborators - show ALL names from collaborator_names field
                        collab_names = task.get('collaborator_names', [])
                        if not collab_names:
                            # Fallback: try to build from collaborators IDs
                            collaborators = task.get('collaborators', []) or []
                            if isinstance(collaborators, list) and collaborators:
                                collab_names = [f'User {c}' for c in collaborators]
                        
                        collab_str = ', '.join(collab_names) if collab_names else 'None'
                        
                        # Duration
                        completion_days = task.get('completion_days')
                        duration_text = f"{completion_days}d" if completion_days else "Ongoing"
                        
                        task_data.append([
                            task_name,
                            task.get('status', 'Unknown'),
                            task.get('priority', 'Normal'),
                            owner_name,
                            collab_str,
                            task.get('due_date', '')[:10] if task.get('due_date') else 'N/A',
                            duration_text
                        ])
                    
                    # Create task table - wider columns for full names
                    task_table = Table(task_data, colWidths=[2.2*inch, 0.7*inch, 0.6*inch, 1.0*inch, 1.2*inch, 0.7*inch, 0.6*inch])
                    
                    # Base style
                    table_style = [
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 7),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP')
                    ]
                    
                    # Add yellow highlighting for member's tasks
                    for row_idx in highlighted_rows:
                        table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightyellow))
                        table_style.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.black))
                    
                    task_table.setStyle(TableStyle(table_style))
                    story.append(task_table)
                    story.append(Spacer(1, 15))

    def export_personal_report_excel(self, report_data: ReportData) -> bytes:
        """Export personal report as Excel with Dashboard and Projects with tasks"""
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # ========== DASHBOARD SHEET ==========
            performance_data = []
            
            # KPI Section
            performance_data.extend([
                ['PERFORMANCE DASHBOARD', ''],
                ['Name', report_data.user_name],
                ['Role', report_data.user_role.capitalize() if report_data.user_role else 'Unknown'],
                ['', ''],
                ['KEY METRICS', ''],
                ['Task Completion Rate', f"{report_data.completion_percentage:.1f}%"],
                ['Tasks Completed', f"{report_data.completed_tasks}/{report_data.total_tasks}"],
                ['Overdue Tasks', f"{getattr(report_data, 'overdue_tasks', 0)} ({getattr(report_data, 'overdue_percentage', 0):.1f}%)"],
                ['Active Projects', str(report_data.total_projects)],
                ['Avg Task Duration', f"{report_data.average_task_duration:.1f} days" if report_data.average_task_duration else "N/A"],
                ['', ''],
                ['TASK STATUS BREAKDOWN', '']
            ])
            
            # Task status breakdown
            for status, count in report_data.task_stats.items():
                performance_data.append([f"{status} Tasks", str(count)])
            
            # Performance indicators
            performance_data.extend([
                ['', ''],
                ['PERFORMANCE INDICATORS', ''],
                ['Overall Status', 
                 '🟢 Excellent' if report_data.completion_percentage >= 80 and getattr(report_data, 'overdue_percentage', 0) < 15 else
                 '🟡 Good' if report_data.completion_percentage >= 60 else '🔴 Needs Improvement'],
                ['Efficiency Rating',
                 '🟢 High' if report_data.average_task_duration and report_data.average_task_duration <= 5 else
                 '🟡 Medium' if report_data.average_task_duration and report_data.average_task_duration <= 10 else '🔴 Low']
            ])
            
            dashboard_df = pd.DataFrame(performance_data, columns=['Metric', 'Value'])
            dashboard_df.to_excel(writer, sheet_name='Dashboard', index=False)

            # ========== PROJECTS WITH TASKS SHEET ==========
            # Get projects breakdown which contains both project stats and task details
            projects_breakdown = getattr(report_data, 'projects_breakdown', [])
            user_id = report_data.user_id
            
            if projects_breakdown:
                all_project_data = []
                
                for project in projects_breakdown:
                    # Project header
                    all_project_data.append([f"PROJECT: {project.get('project_name', 'Unknown Project')}", '', '', '', '', '', '', '', '', ''])
                    all_project_data.append(['', '', '', '', '', '', '', '', '', ''])
                    
                    # Project-level metrics
                    all_project_data.extend([
                        ['Project Metrics', '', '', '', '', '', '', '', '', ''],
                        ['Total Tasks in Project', project.get('total_tasks', 0), '', '', '', '', '', '', '', ''],
                        ['Completed Tasks', project.get('completed_tasks', 0), '', '', '', '', '', '', '', ''],
                        ['Ongoing Tasks', project.get('in_progress_tasks', 0), '', '', '', '', '', '', '', ''],
                        ['Under Review Tasks', project.get('under_review_tasks', 0), '', '', '', '', '', '', '', ''],
                        ['Overdue Tasks', project.get('overdue_tasks', 0), '', '', '', '', '', '', '', ''],
                        ['Completion Rate', f"{project.get('completion_percentage', 0):.1f}%", '', '', '', '', '', '', '', ''],
                        ['Average Task Duration', f"{project.get('average_task_duration'):.1f} days" if project.get('average_task_duration') else "N/A", '', '', '', '', '', '', '', ''],
                        ['Projected Completion', project.get('projected_completion_date') if project.get('projected_completion_date') else "N/A", '', '', '', '', '', '', '', ''],
                        ['', '', '', '', '', '', '', '', '', '']
                    ])
                    
                    # ALL Tasks in this project
                    project_tasks = project.get('task_details', [])
                    if project_tasks:
                        all_project_data.append(['ALL TASKS IN THIS PROJECT', '', '', '', '', '', '', '', '', ''])
                        all_project_data.append(['(★ = You are owner or collaborator)', '', '', '', '', '', '', '', '', ''])
                        all_project_data.append(['', '', '', '', '', '', '', '', '', ''])
                        
                        # Task table headers
                        all_project_data.append([
                            'Task Name',
                            'Status',
                            'Priority',
                            'Owner',
                            'Collaborators',
                            'Created Date',
                            'Due Date',
                            'Completed Date',
                            'Duration',
                            'You Involved'
                        ])
                        
                        # Add all tasks
                        for task in project_tasks:
                            # Check if user is owner or collaborator
                            owner_id = task.get('owner_id')
                            collaborator_ids = task.get('collaborator_ids', []) or []
                            is_user_involved = (owner_id == user_id) or (user_id in collaborator_ids)
                            
                            # Get owner name
                            owner_name = task.get('owner_name', 'Unknown')
                            
                            # Get collaborators - use collaborators field which has names
                            collaborators = task.get('collaborators', []) or []
                            collab_str = ', '.join(collaborators) if collaborators else 'None'
                            
                            # Duration calculation
                            completion_days = task.get('completion_days')
                            duration_text = f"{completion_days} days" if completion_days else "Ongoing"
                            
                            # Mark task name with star if user is involved
                            task_name = task.get('task_name', 'Unknown Task')
                            if is_user_involved:
                                task_name = f"★ {task_name}"
                            
                            all_project_data.append([
                                task_name,
                                task.get('status', 'Unknown'),
                                task.get('priority', 'Normal'),
                                owner_name,
                                collab_str,
                                task.get('created_at', '')[:10] if task.get('created_at') else 'N/A',
                                task.get('due_date', '')[:10] if task.get('due_date') else 'N/A',
                                task.get('completed_at', '')[:10] if task.get('completed_at') else 'N/A',
                                duration_text,
                                '★ YES' if is_user_involved else 'No'
                            ])
                        
                        all_project_data.append(['', '', '', '', '', '', '', '', '', ''])
                        all_project_data.append(['', '', '', '', '', '', '', '', '', ''])
                
                # Create DataFrame and write to Excel
                projects_df = pd.DataFrame(all_project_data)
                projects_df.to_excel(writer, sheet_name='Projects & Tasks', index=False, header=False)
                
                # Apply highlighting to rows where user is involved
                worksheet = writer.sheets['Projects & Tasks']
                from openpyxl.styles import PatternFill
                highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Go through rows and highlight where user is involved
                for row_idx, row_data in enumerate(all_project_data, start=1):
                    if len(row_data) > 0 and isinstance(row_data[0], str):
                        # Check if this is a task row with user involvement
                        if row_data[0].startswith('★'):
                            # Highlight entire row
                            for col_idx in range(1, 11):  # 10 columns
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = highlight_fill

        buffer.seek(0)
        return buffer.getvalue()

    def export_team_report_excel(self, manager_report: ReportData = None, team_report: TeamReportData = None, detailed_workload: Dict[str, Any] = None) -> bytes:
        """Export team report with Overview tab and individual member tabs showing all project tasks with highlighting"""
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # ========== TAB 1: OVERVIEW ==========
            dashboard_data = []
            
            # Team Overview Section
            dashboard_data.extend([
                ['TEAM PERFORMANCE DASHBOARD', ''],
                ['Department' if team_report.dept_name and not team_report.team_name else 'Team', 
                 team_report.dept_name or team_report.team_name or 'Unknown'],
                ['Department Size' if team_report.dept_name and not team_report.team_name else 'Team Size', 
                 len(team_report.member_reports)],
                ['', ''],
                ['TEAM METRICS', ''],
                ['Total Tasks', team_report.total_team_tasks],
                ['Team Completion Rate', f"{team_report.team_completion_percentage:.1f}%"],
                ['Team Overdue Rate', f"{getattr(team_report, 'team_overdue_percentage', 0):.1f}%"],
                ['Active Projects', team_report.total_team_projects],
                ['Avg Task Duration', f"{team_report.team_average_task_duration:.1f} days" if team_report.team_average_task_duration else "N/A"],
                ['', ''],
                ['PERFORMANCE RATING', ''],
                ['Team Status', 
                 '🟢 High Performing' if team_report.team_completion_percentage >= 75 and getattr(team_report, 'team_overdue_percentage', 0) < 20 else
                 '🟡 Moderate Performance' if team_report.team_completion_percentage >= 50 else '🔴 Needs Improvement']
            ])
            
            # Task status breakdown
            team_task_stats = getattr(team_report, 'team_task_stats', {})
            if team_task_stats:
                dashboard_data.extend([['', ''], ['TASK STATUS BREAKDOWN', '']])
                for status, count in team_task_stats.items():
                    dashboard_data.append([f"{status} Tasks", count])
            
            dashboard_df = pd.DataFrame(dashboard_data, columns=['Metric', 'Value'])
            dashboard_df.to_excel(writer, sheet_name='Overview', index=False)
            
            # ========== TEAM MEMBER TABS (One per member) ==========
            # Get enriched project stats from team report
            team_project_stats = getattr(team_report, 'team_project_stats', [])
            
            # Detect if this is a department report
            is_department_report = team_report.dept_name and not team_report.team_name
            
            # Group members by team if this is a department report
            members_to_process = []
            if is_department_report:
                # Group members by team_id
                teams_dict = {}
                for member in team_report.member_reports:
                    team_id = getattr(member, 'team_id', None)
                    team_name = getattr(member, 'team_name', None)
                    
                    # Handle missing team info
                    if not team_id:
                        team_id = 'no_team'
                        team_name = 'No Team Assigned'
                    elif not team_name:
                        # If team_id exists but team_name is None, use fallback
                        team_name = f'Team {team_id}'
                    
                    if team_id not in teams_dict:
                        teams_dict[team_id] = {
                            'team_name': team_name,
                            'members': []
                        }
                    teams_dict[team_id]['members'].append(member)
                
                # Process teams in order, adding team headers
                for team_id, team_data in teams_dict.items():
                    # Add a team header placeholder with member list
                    members_to_process.append({
                        'is_team_header': True,
                        'team_name': team_data['team_name'],
                        'team_members': team_data['members']  # Pass the actual members
                    })
                    # Add team members
                    members_to_process.extend(team_data['members'])
            else:
                # Regular team report - just process members as-is
                members_to_process = team_report.member_reports
            
            # Check if there are no members to process
            if not members_to_process:
                # Add a note sheet
                no_members_df = pd.DataFrame([
                    ['Team/Department Report', ''],
                    ['', ''],
                    ['Status', 'No members in this team/department'],
                    ['', ''],
                    ['Note', 'This team/department currently has no members assigned.']
                ], columns=['Field', 'Value'])
                no_members_df.to_excel(writer, sheet_name='No Members', index=False, header=False)
            
            for member_idx, member in enumerate(members_to_process, 1):
                # Skip team headers in member processing (we'll create a sheet for them)
                if isinstance(member, dict) and member.get('is_team_header'):
                    # Create a team header sheet with member list
                    team_header_sheet = f"TEAM - {member['team_name'][:20]}"
                    team_header_data = [
                        [f"TEAM: {member['team_name']}", ''],
                        ['', ''],
                        ['Team Members:', ''],
                        ['', '']
                    ]
                    
                    # Add list of team members
                    team_members = member.get('team_members', [])
                    for idx, team_member in enumerate(team_members, 1):
                        team_header_data.append([
                            f"{idx}. {team_member.user_name}",
                            team_member.user_role.capitalize() if team_member.user_role else 'Unknown'
                        ])
                    
                    team_df = pd.DataFrame(team_header_data, columns=['Member', 'Role'])
                    team_df.to_excel(writer, sheet_name=team_header_sheet, index=False)
                    continue
                
                sheet_name = f"{member.user_name[:18]}"  # Limit name length for Excel
                
                # Member Info Section
                member_info_data = []
                
                # Add team info for department reports
                if is_department_report:
                    team_name = getattr(member, 'team_name', None)
                    if not team_name:
                        team_id = getattr(member, 'team_id', None)
                        team_name = f'Team {team_id}' if team_id else 'No Team Assigned'
                    
                    member_info_data.extend([
                        [f'MEMBER {member_idx}: {member.user_name}', '', '', '', '', '', '', '', '', ''],
                        ['Team', team_name, '', '', '', '', '', '', '', ''],
                        ['Role', member.user_role.capitalize() if member.user_role else 'Unknown', '', '', '', '', '', '', '', ''],
                    ])
                else:
                    member_info_data.extend([
                        [f'MEMBER {member_idx}: {member.user_name}', '', '', '', '', '', '', '', '', ''],
                        ['Role', member.user_role.capitalize() if member.user_role else 'Unknown', '', '', '', '', '', '', '', ''],
                    ])
                
                member_info_data.extend([
                    ['', '', '', '', '', '', '', '', '', ''],
                    ['MEMBER PERFORMANCE', '', '', '', '', '', '', '', '', ''],
                    ['Total Projects', member.total_projects, '', '', '', '', '', '', '', ''],
                    ['Total Tasks', member.total_tasks, '', '', '', '', '', '', '', ''],
                    ['Completed Tasks', member.completed_tasks, '', '', '', '', '', '', '', ''],
                    ['Completion Rate', f"{member.completion_percentage:.1f}%", '', '', '', '', '', '', '', ''],
                    ['Overdue Tasks', getattr(member, 'overdue_tasks', 0), '', '', '', '', '', '', '', ''],
                    ['Average Task Duration', f"{member.average_task_duration:.1f} days" if member.average_task_duration else "N/A", '', '', '', '', '', '', '', ''],
                    ['', '', '', '', '', '', '', '', '', ''],
                    ['', '', '', '', '', '', '', '', '', '']
                ])
                
                # Projects this member is involved in
                member_user_id = member.user_id
                member_projects = []
                
                for project in team_project_stats:
                    project_id = project.get('project_id')
                    member_involvement = project.get('member_involvement', {})
                    
                    # Check if this member is involved in the project
                    if member_user_id in member_involvement:
                        involved_task_ids = member_involvement[member_user_id]['involved_tasks']
                        
                        # Get all tasks for this project
                        all_project_tasks = project.get('all_tasks', [])
                        
                        # Add project header
                        member_info_data.append([f"PROJECT: {project.get('project_name', 'Unknown Project')}", '', '', '', '', '', '', '', '', ''])
                        member_info_data.append(['', '', '', '', '', '', '', '', '', ''])
                        
                        # Project-level metrics
                        member_info_data.extend([
                            ['Project Metrics', '', '', '', '', '', '', '', '', ''],
                            ['Total Tasks in Project', project.get('total_tasks', 0), '', '', '', '', '', '', '', ''],
                            ['Completed Tasks', project.get('completed_tasks', 0), '', '', '', '', '', '', '', ''],
                            ['Ongoing Tasks', project.get('in_progress_tasks', 0), '', '', '', '', '', '', '', ''],
                            ['Under Review Tasks', project.get('under_review_tasks', 0), '', '', '', '', '', '', '', ''],
                            ['Overdue Tasks', project.get('overdue_tasks', 0), '', '', '', '', '', '', '', ''],
                            ['Completion Rate', f"{project.get('completion_percentage', 0):.1f}%", '', '', '', '', '', '', '', ''],
                            ['Average Task Duration', f"{project.get('average_task_duration'):.1f} days" if project.get('average_task_duration') else "N/A", '', '', '', '', '', '', '', ''],
                            ['Projected Completion', project.get('projected_completion_date') if project.get('projected_completion_date') else "Completion date cannot be projected", '', '', '', '', '', '', '', ''],
                            ['', '', '', '', '', '', '', '', '', '']
                        ])
                        
                        # ALL Tasks in this project (with highlighting)
                        if all_project_tasks:
                            member_info_data.append(['ALL TASKS IN THIS PROJECT', '', '', '', '', '', '', '', '', ''])
                            member_info_data.append(['(★ = This member is owner or collaborator)', '', '', '', '', '', '', '', '', ''])
                            member_info_data.append(['', '', '', '', '', '', '', '', '', ''])
                            
                            # Task table headers
                            member_info_data.append([
                                'Task Name',
                                'Status',
                                'Priority',
                                'Owner',
                                'Collaborators',
                                'Created Date',
                                'Due Date',
                                'Completed Date',
                                'Duration',
                                'Member Involved'
                            ])
                            
                            # Add all tasks
                            for task in all_project_tasks:
                                task_id = task.get('id')
                                is_member_involved = task_id in involved_task_ids
                                
                                # Get owner name - NO TRUNCATION
                                owner_id = task.get('owner_id')
                                owner_name = task.get('owner_name', 'Unknown')
                                
                                # Get collaborators - show ALL names from collaborator_names field
                                collab_names = task.get('collaborator_names', [])
                                if not collab_names:
                                    # Fallback: try to build from collaborators IDs
                                    collaborators = task.get('collaborators', []) or []
                                    if isinstance(collaborators, list) and collaborators:
                                        collab_names = [f'User {c}' for c in collaborators]
                                
                                collab_str = ', '.join(collab_names) if collab_names else 'None'
                                
                                # Duration calculation
                                completion_days = task.get('completion_days')
                                if not completion_days:
                                    # Calculate if needed
                                    created_at = task.get('created_at')
                                    completed_at = task.get('completed_at')
                                    if created_at and completed_at:
                                        try:
                                            from dateutil import parser as dateparser
                                            created_date = dateparser.parse(created_at)
                                            completed_date = dateparser.parse(completed_at)
                                            completion_days = (completed_date - created_date).days
                                        except:
                                            completion_days = None
                                
                                duration_text = f"{completion_days} days" if completion_days else "Ongoing"
                                
                                # Mark task name with star if member is involved
                                task_name = task.get('task_name', 'Unknown Task')
                                if is_member_involved:
                                    task_name = f"★ {task_name}"
                                
                                member_info_data.append([
                                    task_name,
                                    task.get('status', 'Unknown'),
                                    task.get('priority', 'Normal'),
                                    owner_name,
                                    collab_str,
                                    task.get('created_at', '')[:10] if task.get('created_at') else 'N/A',
                                    task.get('due_date', '')[:10] if task.get('due_date') else 'N/A',
                                    task.get('completed_at', '')[:10] if task.get('completed_at') else 'N/A',
                                    duration_text,
                                    '★ YES' if is_member_involved else 'No'
                                ])
                            
                            member_info_data.append(['', '', '', '', '', '', '', '', '', ''])
                            member_info_data.append(['', '', '', '', '', '', '', '', '', ''])
                
                # Create DataFrame and write to Excel
                member_df = pd.DataFrame(member_info_data)
                member_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                # Apply formatting - highlight member involved tasks
                worksheet = writer.sheets[sheet_name]
                from openpyxl.styles import PatternFill, Font
                
                # Yellow fill for highlighted rows
                highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                bold_font = Font(bold=True)
                
                # Go through rows and highlight where member is involved
                for row_idx, row_data in enumerate(member_info_data, start=1):
                    if len(row_data) > 0 and isinstance(row_data[0], str):
                        # Check if this is a task row with member involvement
                        if row_data[0].startswith('★'):
                            # Highlight entire row
                            for col_idx in range(1, 11):  # 10 columns
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = highlight_fill

        buffer.seek(0)
        return buffer.getvalue()
    
    def export_company_report_pdf(self, company_data: Dict[str, Any]) -> bytes:
        """Export company report as PDF organized by departments and teams, showing projects and tasks for each member"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        # Title
        title = "Company-Wide Performance Report"
        story.append(Paragraph(title, self.title_style))
        story.append(Spacer(1, 20))

        # Company Overview
        story.append(Paragraph("Company Overview", self.heading_style))
        
        company_metrics = company_data.get('company_metrics', {})
        overview_data = [
            ['Metric', 'Value'],
            ['Total Departments', str(company_metrics.get('total_departments', 0))],
            ['Total Teams', str(company_metrics.get('total_teams', 0))],
            ['Total Members', str(company_metrics.get('total_members', 0))],
            ['Total Projects', str(company_metrics.get('total_projects', 0))],
            ['Total Tasks', str(company_metrics.get('total_tasks', 0))],
            ['Completed Tasks', str(company_metrics.get('completed_tasks', 0))],
            ['Company Completion Rate', f"{company_metrics.get('completion_percentage', 0):.1f}%"],
            ['Overdue Tasks', str(company_metrics.get('overdue_tasks', 0))],
            ['Company Overdue Rate', f"{company_metrics.get('overdue_percentage', 0):.1f}%"],
            ['Average Task Duration', f"{company_metrics.get('average_task_duration'):.1f} days" if company_metrics.get('average_task_duration') else "N/A"]
        ]
        
        overview_table = Table(overview_data, colWidths=[2.5*inch, 2.5*inch])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(overview_table)
        story.append(Spacer(1, 30))

        # Department sections
        departments = company_data.get('departments', [])
        
        for dept_idx, department in enumerate(departments):
            if dept_idx > 0:
                story.append(PageBreak())
            
            dept_name = department.get('dept_name', 'Unknown Department')
            
            # Department Header
            dept_header = Paragraph(f"DEPARTMENT: {dept_name}", 
                                   ParagraphStyle('DeptHeader',
                                                parent=getSampleStyleSheet()['Heading1'],
                                                fontSize=18,
                                                textColor=colors.darkred,
                                                spaceAfter=20))
            story.append(dept_header)
            
            # Director Section with Projects and Tasks
            director_data = department.get('director')
            if director_data:
                story.append(Paragraph(f"Director: {director_data.get('user_name', 'Unknown')}", self.heading_style))
                
                director_summary = [
                    ['Metric', 'Value'],
                    ['Total Projects', str(director_data.get('total_projects', 0))],
                    ['Total Tasks', str(director_data.get('total_tasks', 0))],
                    ['Completed Tasks', str(director_data.get('completed_tasks', 0))],
                    ['Completion Rate', f"{director_data.get('completion_percentage', 0):.1f}%"],
                    ['Overdue Tasks', str(director_data.get('overdue_tasks', 0))],
                    ['Average Task Duration', f"{director_data.get('average_task_duration'):.1f} days" if director_data.get('average_task_duration') else "N/A"]
                ]
                
                director_table = Table(director_summary, colWidths=[2.5*inch, 2*inch])
                director_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(director_table)
                story.append(Spacer(1, 15))
                
                # Director's Projects and Tasks
                self._add_member_projects_and_tasks_to_pdf(story, director_data, director_data.get('user_id'))
            
            # Teams in this department
            teams = department.get('teams', [])
            
            if not teams:
                # No teams in this department
                if director_data:
                    story.append(Spacer(1, 20))
                story.append(Paragraph("No teams in this department", 
                                     ParagraphStyle('NoTeams',
                                                  parent=getSampleStyleSheet()['Normal'],
                                                  fontSize=11,
                                                  textColor=colors.grey,
                                                  spaceAfter=15)))
            
            for team_idx, team in enumerate(teams):
                if team_idx > 0 or director_data:
                    story.append(PageBreak())
                
                team_name = team.get('team_name', 'Unknown Team')
                
                # Team Header
                team_header = Paragraph(f"TEAM: {team_name}", 
                                       ParagraphStyle('TeamHeader',
                                                    parent=getSampleStyleSheet()['Heading2'],
                                                    fontSize=14,
                                                    textColor=colors.darkgreen,
                                                    spaceAfter=15))
                story.append(team_header)
                
                # Manager Section with Projects and Tasks
                manager_data = team.get('manager')
                if manager_data:
                    story.append(Paragraph(f"Manager: {manager_data.get('user_name', 'Unknown')}", 
                                         ParagraphStyle('ManagerHeader',
                                                      parent=getSampleStyleSheet()['Heading3'],
                                                      fontSize=11,
                                                      spaceAfter=10)))
                    
                    manager_summary = [
                        ['Metric', 'Value'],
                        ['Total Projects', str(manager_data.get('total_projects', 0))],
                        ['Total Tasks', str(manager_data.get('total_tasks', 0))],
                        ['Completed Tasks', str(manager_data.get('completed_tasks', 0))],
                        ['Completion Rate', f"{manager_data.get('completion_percentage', 0):.1f}%"],
                        ['Overdue Tasks', str(manager_data.get('overdue_tasks', 0))]
                    ]
                    
                    manager_table = Table(manager_summary, colWidths=[2*inch, 1.8*inch])
                    manager_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(manager_table)
                    story.append(Spacer(1, 15))
                    
                    # Manager's Projects and Tasks
                    self._add_member_projects_and_tasks_to_pdf(story, manager_data, manager_data.get('user_id'))
                else:
                    story.append(Paragraph("No manager assigned to this team", 
                                         ParagraphStyle('NoManager',
                                                      parent=getSampleStyleSheet()['Normal'],
                                                      fontSize=10,
                                                      textColor=colors.grey,
                                                      spaceAfter=15)))
                
                # Staff Members with Projects and Tasks
                staff_list = team.get('staff', [])
                
                if not staff_list:
                    story.append(Paragraph("No staff members in this team", 
                                         ParagraphStyle('NoStaff',
                                                      parent=getSampleStyleSheet()['Normal'],
                                                      fontSize=10,
                                                      textColor=colors.grey,
                                                      spaceAfter=15)))
                
                for staff_idx, staff in enumerate(staff_list):
                    story.append(PageBreak())
                    
                    story.append(Paragraph(f"Staff: {staff.get('user_name', 'Unknown')}", 
                                         ParagraphStyle('StaffHeader',
                                                      parent=getSampleStyleSheet()['Heading3'],
                                                      fontSize=11,
                                                      spaceAfter=10)))
                    
                    staff_summary = [
                        ['Metric', 'Value'],
                        ['Total Projects', str(staff.get('total_projects', 0))],
                        ['Total Tasks', str(staff.get('total_tasks', 0))],
                        ['Completed Tasks', str(staff.get('completed_tasks', 0))],
                        ['Completion Rate', f"{staff.get('completion_percentage', 0):.1f}%"],
                        ['Overdue Tasks', str(staff.get('overdue_tasks', 0))]
                    ]
                    
                    staff_table = Table(staff_summary, colWidths=[2*inch, 1.8*inch])
                    staff_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightyellow),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(staff_table)
                    story.append(Spacer(1, 15))
                    
                    # Staff's Projects and Tasks
                    self._add_member_projects_and_tasks_to_pdf(story, staff, staff.get('user_id'))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def _add_member_projects_and_tasks_to_pdf(self, story, member_data, member_user_id):
        """Add projects and tasks breakdown for a member to the PDF story"""
        projects_breakdown = member_data.get('projects_breakdown', [])
        
        if not projects_breakdown:
            story.append(Paragraph("No projects assigned", self.styles['Normal']))
            story.append(Spacer(1, 15))
            return
        
        for project in projects_breakdown:
            project_name = project.get('project_name', 'Unknown Project')
            
            # Project Header
            story.append(Paragraph(f"Project: {project_name}", 
                                 ParagraphStyle('ProjectTitle',
                                              parent=getSampleStyleSheet()['Heading4'],
                                              fontSize=10,
                                              textColor=colors.darkblue,
                                              spaceAfter=8)))
            
            # Project Summary
            project_summary = [
                ['Project Metric', 'Value'],
                ['Total Tasks', str(project.get('total_tasks', 0))],
                ['Completed Tasks', str(project.get('completed_tasks', 0))],
                ['Completion Rate', f"{project.get('completion_percentage', 0):.1f}%"],
                ['Overdue Tasks', str(project.get('overdue_tasks', 0))],
                ['Average Duration', f"{project.get('average_task_duration', 0):.1f} days" if project.get('average_task_duration') else "N/A"],
                ['Projected Completion', project.get('projected_completion_date', 'N/A')]
            ]
            
            project_table = Table(project_summary, colWidths=[1.8*inch, 2*inch])
            project_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(project_table)
            story.append(Spacer(1, 10))
            
            # Tasks in this project
            task_details = project.get('task_details', [])
            if task_details:
                story.append(Paragraph(f"Tasks ({len(task_details)}) - ★ = Member is owner/collaborator", 
                                     ParagraphStyle('TasksSubtitle',
                                                  parent=getSampleStyleSheet()['Normal'],
                                                  fontSize=8,
                                                  spaceAfter=6)))
                
                task_data = [['Task Name', 'Status', 'Owner', 'Collaborators', 'Due Date', 'Duration']]
                
                # Track which rows need highlighting
                highlighted_rows = []
                
                for task_idx, task in enumerate(task_details):
                    # Check if member is owner or collaborator
                    owner_id = task.get('owner_id')
                    collaborator_ids = task.get('collaborator_ids', []) or []
                    
                    is_member_involved = (owner_id == member_user_id) or (member_user_id in collaborator_ids)
                    
                    # Get task information
                    task_name = task.get('task_name', 'Unknown')
                    if is_member_involved:
                        task_name = f"★ {task_name}"
                        highlighted_rows.append(task_idx + 1)  # +1 because row 0 is header
                    
                    status = task.get('status', 'Unknown')
                    owner = task.get('owner_name', 'Unknown')
                    collaborators_display = ", ".join(task.get('collaborators', [])) if task.get('collaborators') else "None"
                    due_date = task.get('due_date', '')[:10] if task.get('due_date') else 'N/A'
                    
                    # Duration
                    completion_days = task.get('completion_days')
                    duration_text = f"{completion_days}d" if completion_days else "Ongoing"
                    
                    task_data.append([
                        task_name,
                        status,
                        owner,
                        collaborators_display,
                        due_date,
                        duration_text
                    ])
                
                task_table = Table(task_data, colWidths=[1.8*inch, 0.7*inch, 0.9*inch, 1.2*inch, 0.7*inch, 0.6*inch])
                
                # Build table style with highlighting for member's tasks
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 7),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP')
                ]
                
                # Add yellow highlighting for rows where member is involved
                for row_idx in highlighted_rows:
                    table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightyellow))
                
                task_table.setStyle(TableStyle(table_style))
                story.append(task_table)
                story.append(Spacer(1, 15))

    def export_company_report_excel(self, company_data: Dict[str, Any]) -> bytes:
        """Export company report as Excel with overview and detailed member tabs showing projects and tasks"""
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # ========== COMPANY OVERVIEW TAB ==========
            overview_data = []
            
            company_metrics = company_data.get('company_metrics', {})
            overview_data.extend([
                ['COMPANY PERFORMANCE DASHBOARD', ''],
                ['', ''],
                ['COMPANY METRICS', ''],
                ['Total Departments', company_metrics.get('total_departments', 0)],
                ['Total Teams', company_metrics.get('total_teams', 0)],
                ['Total Members', company_metrics.get('total_members', 0)],
                ['Total Projects', company_metrics.get('total_projects', 0)],
                ['Total Tasks', company_metrics.get('total_tasks', 0)],
                ['Completed Tasks', company_metrics.get('completed_tasks', 0)],
                ['Company Completion Rate', f"{company_metrics.get('completion_percentage', 0):.1f}%"],
                ['Overdue Tasks', company_metrics.get('overdue_tasks', 0)],
                ['Company Overdue Rate', f"{company_metrics.get('overdue_percentage', 0):.1f}%"],
                ['Average Task Duration', f"{company_metrics.get('average_task_duration'):.1f} days" if company_metrics.get('average_task_duration') else "N/A"],
                ['', ''],
                ['PERFORMANCE RATING', ''],
                ['Company Status', 
                 '🟢 High Performing' if company_metrics.get('completion_percentage', 0) >= 75 and company_metrics.get('overdue_percentage', 0) < 20 else
                 '🟡 Moderate Performance' if company_metrics.get('completion_percentage', 0) >= 50 else '🔴 Needs Improvement']
            ])
            
            overview_df = pd.DataFrame(overview_data, columns=['Metric', 'Value'])
            overview_df.to_excel(writer, sheet_name='Company Overview', index=False)

            # ========== MEMBER TABS WITH PROJECTS AND TASKS ==========
            departments = company_data.get('departments', [])
            member_counter = 1
            
            for dept in departments:
                dept_name = dept.get('dept_name', 'Unknown Department')
                
                # Process Director
                director_data = dept.get('director')
                if director_data:
                    sheet_name = f"{member_counter}. {director_data.get('user_name', 'Unknown')}"[:31]  # Excel limit
                    self._add_member_excel_sheet(writer, sheet_name, director_data, dept_name, None, 'Director')
                    member_counter += 1
                
                # Process Teams
                teams = dept.get('teams', [])
                
                if not teams:
                    # Add a sheet noting no teams in this department
                    sheet_name = f"{dept_name[:25]} - Info"[:31]
                    no_teams_df = pd.DataFrame([
                        [f"DEPARTMENT: {dept_name}", ''],
                        ['', ''],
                        ['Status', 'No teams in this department'],
                        ['', ''],
                        ['Note', 'This department currently has no teams assigned.']
                    ], columns=['Field', 'Value'])
                    no_teams_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                for team in teams:
                    team_name = team.get('team_name', 'Unknown Team')
                    
                    # Process Manager
                    manager_data = team.get('manager')
                    if manager_data:
                        sheet_name = f"{member_counter}. {manager_data.get('user_name', 'Unknown')}"[:31]
                        self._add_member_excel_sheet(writer, sheet_name, manager_data, dept_name, team_name, 'Manager')
                        member_counter += 1
                    
                    # Process Staff
                    staff_list = team.get('staff', [])
                    
                    if not staff_list and manager_data:
                        # Add note that there are no staff members
                        # This information can be added to the overview or we can skip it
                        # For now, the manager's sheet will show this implicitly
                        pass
                    
                    for staff in staff_list:
                        sheet_name = f"{member_counter}. {staff.get('user_name', 'Unknown')}"[:31]
                        self._add_member_excel_sheet(writer, sheet_name, staff, dept_name, team_name, 'Staff')
                        member_counter += 1

        buffer.seek(0)
        return buffer.getvalue()

    def _add_member_excel_sheet(self, writer, sheet_name, member_data, dept_name, team_name, role):
        """Add a sheet for a member with their projects and tasks"""
        member_rows = []
        
        # Member Header
        member_rows.extend([
            [f"{role.upper()}: {member_data.get('user_name', 'Unknown')}", '', '', '', '', '', ''],
            ['Department', dept_name, '', '', '', '', ''],
        ])
        
        if team_name:
            member_rows.append(['Team', team_name, '', '', '', '', ''])
        
        member_rows.extend([
            ['', '', '', '', '', '', ''],
            ['SUMMARY', '', '', '', '', '', ''],
            ['Total Projects', member_data.get('total_projects', 0), '', '', '', '', ''],
            ['Total Tasks', member_data.get('total_tasks', 0), '', '', '', '', ''],
            ['Completed Tasks', member_data.get('completed_tasks', 0), '', '', '', '', ''],
            ['Completion Rate', f"{member_data.get('completion_percentage', 0):.1f}%", '', '', '', '', ''],
            ['Overdue Tasks', member_data.get('overdue_tasks', 0), '', '', '', '', ''],
            ['Average Duration', f"{member_data.get('average_task_duration'):.1f} days" if member_data.get('average_task_duration') else "N/A", '', '', '', '', ''],
            ['', '', '', '', '', '', '']
        ])
        
        # Projects and Tasks
        projects_breakdown = member_data.get('projects_breakdown', [])
        member_user_id = member_data.get('user_id')
        
        if projects_breakdown:
            for project in projects_breakdown:
                project_name = project.get('project_name', 'Unknown Project')
                
                # Project Header
                member_rows.extend([
                    [f"PROJECT: {project_name}", '', '', '', '', '', ''],
                    ['Total Tasks', project.get('total_tasks', 0), '', '', '', '', ''],
                    ['Completed', project.get('completed_tasks', 0), '', '', '', '', ''],
                    ['Completion Rate', f"{project.get('completion_percentage', 0):.1f}%", '', '', '', '', ''],
                    ['Overdue', project.get('overdue_tasks', 0), '', '', '', '', ''],
                    ['Avg Duration', f"{project.get('average_task_duration'):.1f} days" if project.get('average_task_duration') else "N/A", '', '', '', '', ''],
                    ['Projected Completion', project.get('projected_completion_date', 'N/A'), '', '', '', '', ''],
                    ['', '', '', '', '', '', ''],
                    ['TASKS (★ = Member is owner/collaborator)', '', '', '', '', '', ''],
                    ['Task Name', 'Status', 'Priority', 'Owner', 'Collaborators', 'Due Date', 'Duration']
                ])
                
                # Tasks
                task_details = project.get('task_details', [])
                for task in task_details:
                    # Check if member is owner or collaborator
                    owner_id = task.get('owner_id')
                    collaborator_ids = task.get('collaborator_ids', []) or []
                    is_member_involved = (owner_id == member_user_id) or (member_user_id in collaborator_ids)
                    
                    task_name = task.get('task_name', 'Unknown')
                    if is_member_involved:
                        task_name = f"★ {task_name}"
                    
                    collaborators_display = ", ".join(task.get('collaborators', [])) if task.get('collaborators') else "None"
                    completion_days = task.get('completion_days')
                    duration_text = f"{completion_days}d" if completion_days else "Ongoing"
                    
                    member_rows.append([
                        task_name,
                        task.get('status', 'Unknown'),
                        task.get('priority', 'Normal'),
                        task.get('owner_name', 'Unknown'),
                        collaborators_display,
                        task.get('due_date', '')[:10] if task.get('due_date') else 'N/A',
                        duration_text
                    ])
                
                member_rows.extend([
                    ['', '', '', '', '', '', ''],
                    ['', '', '', '', '', '', '']
                ])
        else:
            member_rows.append(['No projects assigned', '', '', '', '', '', ''])
        
        # Create DataFrame and write to Excel
        member_df = pd.DataFrame(member_rows)
        member_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        # Apply formatting
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Highlight member's tasks (rows with ★)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        for idx, row in enumerate(member_rows, start=1):
            if row and isinstance(row[0], str) and row[0].startswith('★'):
                for col in range(1, 8):
                    worksheet.cell(row=idx, column=col).fill = yellow_fill